'''
import numpy as np
x=[1,2,3,4,5]
y=np.array(x)
print(y)
-----------------------------------------------
from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A3'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")
--------------------------------------------------

from openpyxl import Workbook
wb = Workbook()

ws=wb.active
ws['A1'] = 42
ws.append([1, 2, 3])

import datetime
ws['A3'] = datetime.datetime.now()

wb.save('sample.xlsx')
------------------------------------------------------

from openpyxl import Workbook
wb=Workbook()

ws=wb.active
ws['A1'] = 42
ws.append([1, 2, 3])

import datetime
ws['A3'] = datetime.datetime.now()

wb.save('sample.xlsx')
---------------------------------------------------

from openpyxl import Workbook
wb = Workbook()

ws = wb.active
ws['A1'] = 25
ws.append([1,2,3])
import datetime
ws['A3'] = datetime.datetime.now()

wb.save('sample.xlsx')
-----------------------------------------------------

from openpyxl import Workbook
from openpyxl.chart import (
    AreaChart,
    Reference,
    Series,
    )
wb = Workbook()
ws = wb.active

rows = [
    ['Number', 'Batch 1', 'Batch 2'],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 10],
    [6, 25, 5],
    [7, 50, 10],
    ]
for row in rows:
    ws.append(row)

chart = AreaChart()
chart.title = "Area Chart"
chart.style = 48
chart.x_axis.title = 'Test'
chart.y_axis.title = 'Percentage'

cats = Reference(ws, min_col=1, min_row=1, max_row=15)
data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=7)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

ws.add_chart(chart, "A10")

wb.save("area.xlsx")

--------------------------------------------------------

import csv
with open("movie.csv", "r") as f:
    reader = csv.reader(f)
    print(reader,type(reader))
    for row in reader:
        print (row[1])


------------------------------------------------------------
'''
import csv
with open('movie.csv','r') as f:
    reader=csv.reader(f)
    print(reader,type(reader))
    for row in reader:
        print(row[0])
'''
------------------------------------------------------------

import csv
with open('movie.csv','r') as f:
    reader=csv.reader(f)
    print(reader,type(reader))
    for row in reader:
        print(row[2])
------------------------------------------------------

import csv
with open('movie.csv','r') as f:
    reader=csv.reader(f)
    print(reader,type(reader))
    for row in reader:
        print(row[0])
-------------------------------------------------------

import csv
with open('movie.csv','r') as f:
    reader=csv.reader(f,delimiter=":")
    print(reader,type(reader))
    for row in reader:
        print(row[0])

---------------------------------------------------------

import csv
with open('x.csv','r') as f:
    reader=csv.reader(f)
    print(reader,type(reader))
    for row in reader:
        print(row[0])
        
------------------------------------------------

import csv
with open('movie.csv','r') as f:
    reader=csv.DictReader(f)
    print(reader,type(reader))
    for row in reader:
        print(row['Use'])

        if row['Use'] == 'Retail':
                total += float(row['Square Footage'])
        print("Retail space = {}".format(total))
     
---------------------------------------------------

import csv
with open('filename', "w") as csvfile:
        writer = csv.writer(csvfile,delimiter=',')
        #for point in writer:
        writer.writerow([1])
        print( "Points written sucessfully to file")
       
----------------------------------------------------------

import json
with open('person.json') as f:
    data=json.load(f)
    print(data)

----------------------------------------------------------

import json
with open('person.json') as f:
    data=json.load(f)
    print(data)
-------------------------------------------------------

import json
with open('person.json') as f:
    data=json.load(f)
    print(data)

-----------------------------------------------------

import json
person_dict={'name':'bob','age':12,'children':None}
person_json=json.dumps(person_dict)
print(person_json)

----------------------------------------------------------

import json
person_dict={'name':'Bob','languages':['english','french'],
             'married':True,'age':32}
with open('person1.txt','w') as json_file:
    json.dump(person_dict,json_file)
    
-------------------------------------------------------

import json
person_string='{"name":"Bob","languages":"English",\
"numbers":[2,1.6,null]}'
person_dict=json.loads(person_string)
print(json.dumps(person_dict,indent=4,sort_keys=True))
'''






























































































































































