'''
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = Workbook()
ws = wb.active

data = [
    ['Apples', 10000, 5000, 8000, 6000],
    ['Pears',   2000, 3000, 4000, 5000],
    ['Bananas', 6000, 6000, 6500, 6000],
    ['Oranges',  500,  300,  200,  700],
]

# add column headings. NB. these must be strings
ws.append(["Fruit", "2011", "2012", "2013", "2014"])
for row in data:
    ws.append(row)

tab = Table(displayName="Table1", ref="A1:E5")

#Add a default style with striped rows and banded columns
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab.tableStyleInfo = style
ws.add_table(tab)
wb.save("table.xlsx")

-------------------------------------------------------
#Writing to a cell

# import load_workbook
from openpyxl import load_workbook
# set file path
filepath="table.xlsx"
# load demo.xlsx 
wb=load_workbook(filepath)
# select demo.xlsx
sheet=wb.active
# set value for cell A1=1
sheet['A1'] = 1
# set value for cell B2=2
sheet.cell(row=2, column=2).value = 2
# save workbook 
wb.save(filepath)
--------------------------------------------------
#Appending group of values at the bottom of the current sheet

# import Workbook
from openpyxl import Workbook
# create Workbook object
wb=Workbook()
# set file path
filepath="demo.xlsx"
# select demo.xlsx
sheet=wb.active
data=[('Id','Name','Marks'),
      (1,'ABC',50),
      (2,'CDE',100)]
# append all rows
for row in data:
    sheet.append(row)
# save file
wb.save(filepath)
----------------------------------------------
#Reading a cell

# import load_workbook
from openpyxl import load_workbook
# set file path
filepath="demo.xlsx"
# load demo.xlsx 
wb=load_workbook(filepath)
# select demo.xlsx
sheet=wb.active
# get b1 cell value
b1=sheet['B1']
# get b2 cell value
b2=sheet['B2']
# get b3 cell value
b3=sheet.cell(row=3,column=2)
# print b1, b2 and b3
print(b1)
print(b2)
print(b3)
------------------------------------------
#Iterating by rows

# import load_workbook
from openpyxl import load_workbook
# set file path
filepath="demo.xlsx"
# load demo.xlsx 
wb=load_workbook(filepath)
# select demo.xlsx
sheet=wb.active
# get max row count
max_row=sheet.max_row
# get max column count
max_column=sheet.max_column
# iterate over all cells 
# iterate over all rows
for i in range(1,max_row+1):
     
     # iterate over all columns
     for j in range(1,max_column+1):
          # get particular cell value    
          cell_obj=sheet.cell(row=i,column=j)
          # print cell value     
          print(cell_obj.value,end=' | ')
     # print new line
     print('\n')
-------------------------------------------------
#Add sheet to the existing xlsx

# import load_workbook
from openpyxl import load_workbook
# set file path
filepath="demo.xlsx"
# load demo.xlsx 
wb=load_workbook(filepath)
# create new sheet
wb.create_sheet('Sheet 2')
# save workbook
wb.save(filepath)
--------------------------------------------------

#Copy data from one sheet to another sheet

# import load_workbook
from openpyxl import load_workbook
# set file path
filepath="demo.xlsx"
# load demo.xlsx 
wb=load_workbook(filepath)
# get Sheet
source=wb.get_sheet_by_name('Sheet')
# copy sheet
target=wb.copy_worksheet(source)
# save workbook
wb.save(filepath)
--------------------------------------------------

#Remove sheet from existing xlsx

# import load_workbook
from openpyxl import load_workbook
# set file path
filepath="demo.xlsx"
# load demo.xlsx 
wb=load_workbook(filepath)
# create new sheet
wb.remove(wb.get_sheet_by_name('Sheet 2'))
# save workbook
wb.save(filepath)
------------------------------------------------
'''

# import pandas lib as pd 
import pandas as pd 
  
# read by default 1st sheet of an excel file 
dataframe1 = pd.read_excel('demo.xlsx') 
  
print(dataframe1)

#-----------------------------------------------------
'''
# importing the csv module 
import csv 
  
# my data rows as dictionary objects 
mydict =[{'branch': 'COE', 'cgpa': '9.0', 'name': 'Nikhil', 'year': '2'}, 
         {'branch': 'COE', 'cgpa': '9.1', 'name': 'Sanchit', 'year': '2'}, 
         {'branch': 'IT', 'cgpa': '9.3', 'name': 'Aditya', 'year': '2'}, 
         {'branch': 'SE', 'cgpa': '9.5', 'name': 'Sagar', 'year': '1'}, 
         {'branch': 'MCE', 'cgpa': '7.8', 'name': 'Prateek', 'year': '3'}, 
         {'branch': 'EP', 'cgpa': '9.1', 'name': 'Sahil', 'year': '2'}] 
  
# field names 
fields = ['name', 'branch', 'year', 'cgpa'] 
  
# name of csv file 
filename = "university_records.csv"
  
# writing to csv file 
with open(filename, 'w') as csvfile: 
    # creating a csv dict writer object 
    writer = csv.DictWriter(csvfile, fieldnames = fields) 
      
    # writing headers (field names) 
    writer.writeheader() 
      
    # writing data rows 
    writer.writerows(mydict)
   
---------------------------------------------- 


# import pandas as pd 
import pandas as pd 
  
# Create a Pandas dataframe from some data. 
df = pd.DataFrame({'Data': ['Geeks', 'For', 'geeks', 'is', 
                               'portal', 'for', 'geeks']}) 
  
# Create a Pandas Excel writer 
# object using XlsxWriter as the engine. 
writer = pd.ExcelWriter('pandasEx.xlsx',  
                   engine ='xlsxwriter') 
  
# Write a dataframe to the worksheet. 
df.to_excel(writer, sheet_name ='Sheet1') 
  
# Close the Pandas Excel writer 
# object and output the Excel file. 
writer.save()
----------------------------------------------------

import pandas as pd
excel_file = 'demo.xls'
movies = pd.read_excel(excel_file)
movies.head()

'''

























































































































